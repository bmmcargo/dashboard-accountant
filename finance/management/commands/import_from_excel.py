from django.core.management.base import BaseCommand
import os
from datetime import datetime
from decimal import Decimal
from openpyxl import load_workbook
from finance.models import InboundTransaction, OutboundTransaction, Penerimaan, Jurnal, Akun, Manifest

class Command(BaseCommand):
    help = 'Import data dari file Excel (Inbound, Outbound, Penerimaan, Manifest)'

    def handle(self, *args, **kwargs):
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
        
        # 1. IMPORT INBOUND dari Excel
        path_inbound = os.path.join(base_dir, 'INBOUND_2026 (1).xlsx')
        if os.path.exists(path_inbound):
            self.import_inbound_excel(path_inbound)
        else:
            self.stdout.write(self.style.WARNING(f'File tidak ditemukan: {path_inbound}'))

        # 2. IMPORT OUTBOUND dari Excel
        path_outbound = os.path.join(base_dir, 'Lap_outbond_bmm_Agustus_2022.xlsx')
        if os.path.exists(path_outbound):
            self.import_outbound_excel(path_outbound)
        else:
            self.stdout.write(self.style.WARNING(f'File tidak ditemukan: {path_outbound}'))

        # 3. IMPORT PENERIMAAN dari CSV (format sudah sederhana)
        path_penerimaan = os.path.join(base_dir, 'Lap_outbond_bmm_penerimaan_diluar_piutang.csv')
        if os.path.exists(path_penerimaan):
            self.import_penerimaan_csv(path_penerimaan)
        else:
            self.stdout.write(self.style.WARNING(f'File tidak ditemukan: {path_penerimaan}'))

        # 4. IMPORT MANIFEST dari CSV files
        self.import_manifest_csvs(base_dir)

    # ==========================
    # IMPORT INBOUND dari Excel
    # ==========================
    def import_inbound_excel(self, filepath):
        self.stdout.write(f"\n📥 Import INBOUND dari Excel: {filepath}")
        
        # Pastikan akun pendapatan inbound ada
        akun_pendapatan, _ = Akun.objects.get_or_create(
            kode='402', 
            defaults={'nama': 'Pendapatan Jasa Inbound', 'kategori': 'REVENUE'}
        )
        akun_kas, _ = Akun.objects.get_or_create(
            kode='101',
            defaults={'nama': 'Kas', 'kategori': 'ASSET'}
        )
        
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        
        count = 0
        jurnal_count = 0
        
        # Header di baris 1, data mulai baris 2
        # Kolom: A=TGL STT, B=TGL MASUK, C=VENDOR, D=NO RESI, E=TUJUAN, F=KOLI, G=KILO, 
        #        H=TGL DOORING, I=TGL KEMBALI, J=KETERANGAN, K=TARIF/KG, L=TOTAL
        for row in ws.iter_rows(min_row=2, values_only=True):
            no_resi = str(row[3]).strip() if row[3] else ''
            if not no_resi or no_resi == 'None':
                continue
            
            try:
                tgl_stt = self.parse_excel_date(row[0])
                tgl_masuk = self.parse_excel_date(row[1])
                vendor = str(row[2]).strip() if row[2] else ''
                tujuan = str(row[4]).strip() if row[4] else ''
                koli = self.parse_int(row[5])
                kilo = self.parse_float(row[6])
                tgl_dooring = self.parse_excel_date(row[7])
                tgl_kembali = self.parse_excel_date(row[8])
                keterangan = str(row[9]).strip() if row[9] else ''
                tarif_kg = str(row[10]).strip() if row[10] else ''
                total = self.parse_rupiah(row[11])
                
                obj, created = InboundTransaction.objects.update_or_create(
                    no_resi=no_resi,
                    defaults={
                        'tanggal_stt': tgl_stt,
                        'tanggal_masuk_stt': tgl_masuk,
                        'vendor': vendor,
                        'tujuan': tujuan,
                        'koli': koli,
                        'kilo': kilo,
                        'tanggal_dooring': tgl_dooring,
                        'tanggal_kembali': tgl_kembali,
                        'keterangan': keterangan,
                        'tarif_per_kg': tarif_kg,
                        'total_biaya': total,
                    }
                )
                count += 1
                
                # Buat Jurnal Pendapatan otomatis
                if total > 0 and tgl_masuk:
                    uraian = f"Pendapatan Inbound: {no_resi} - {vendor} ke {tujuan}"
                    if not Jurnal.objects.filter(tanggal=tgl_masuk, uraian=uraian[:255], nominal=total).exists():
                        Jurnal.objects.create(
                            tanggal=tgl_masuk,
                            uraian=uraian[:255],
                            akun_debit=akun_kas,
                            akun_kredit=akun_pendapatan,
                            nominal=total
                        )
                        jurnal_count += 1
                        
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"  Error Resi {no_resi}: {e}"))
        
        wb.close()
        self.stdout.write(self.style.SUCCESS(f"  ✅ Import {count} Inbound + {jurnal_count} Jurnal Pendapatan"))

    # ==========================
    # IMPORT OUTBOUND dari Excel (dengan merged cells)
    # ==========================
    def import_outbound_excel(self, filepath):
        self.stdout.write(f"\n📤 Import OUTBOUND dari Excel: {filepath}")
        
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        
        count = 0
        
        # Berdasarkan gambar: Header multi-row, data mulai baris 7
        # Kolom (0-indexed setelah iter_rows):
        # A=NO, B=TGL, C=PENGIRIM, D=PENERIMA, E=NO HP, F=No resi BMM, G=Koli, H=KG
        # I=Tarif (Harga BMM), J=Total (Harga BMM)
        # K=Tgl V1, L=No resi V1, M=Biaya V1
        # N=Tgl V2, O=No Resi V2, P=Biaya V2
        # Q=Tgl V3, R=No Resi V3, S=Biaya V3
        # T=Ket, U=Tgl Status, V=Nama Status, W=Profit
        
        for row in ws.iter_rows(min_row=7, values_only=True):
            if not row or len(row) < 6:
                continue
            
            no_resi = str(row[5]).strip() if row[5] else ''
            if not no_resi or no_resi == 'None' or no_resi == '':
                continue
            
            try:
                tanggal = self.parse_excel_date(row[1])
                pengirim = str(row[2]).strip() if row[2] else ''
                penerima = str(row[3]).strip() if row[3] else ''
                no_hp = str(row[4]).strip() if row[4] else ''
                koli = self.parse_int(row[6])
                kg = str(row[7]).strip() if row[7] else ''
                tarif = str(row[8]).strip() if len(row) > 8 and row[8] else ''
                total = self.parse_rupiah(row[9]) if len(row) > 9 else 0
                
                # Vendor 1
                v1_tgl = self.parse_excel_date(row[10]) if len(row) > 10 else None
                v1_resi = str(row[11]).strip() if len(row) > 11 and row[11] else ''
                v1_biaya = self.parse_rupiah(row[12]) if len(row) > 12 else 0
                
                # Vendor 2
                v2_tgl = self.parse_excel_date(row[13]) if len(row) > 13 else None
                v2_resi = str(row[14]).strip() if len(row) > 14 and row[14] else ''
                v2_biaya = self.parse_rupiah(row[15]) if len(row) > 15 else 0
                
                # Keterangan, Status, Profit
                keterangan = str(row[19]).strip() if len(row) > 19 and row[19] else ''
                tgl_bayar = self.parse_excel_date(row[20]) if len(row) > 20 else None
                nama_bayar = str(row[21]).strip() if len(row) > 21 and row[21] else ''
                profit = self.parse_rupiah(row[22]) if len(row) > 22 else 0
                
                # Status bisa di kolom berbeda - cek ket
                status = ''
                if 'COD' in keterangan.upper():
                    status = 'COD'
                elif 'CASH' in keterangan.upper():
                    status = 'CASH'
                elif 'TRANSFER' in keterangan.upper():
                    status = 'TRANSFER'
                
                obj, created = OutboundTransaction.objects.update_or_create(
                    no_resi_bmm=no_resi,
                    defaults={
                        'tanggal': tanggal,
                        'pengirim': pengirim,
                        'penerima': penerima,
                        'no_hp': no_hp,
                        'koli': koli,
                        'kg': kg,
                        'tarif': tarif,
                        'total': total,
                        'vendor1_tgl': v1_tgl,
                        'vendor1_resi': v1_resi,
                        'vendor1_biaya': v1_biaya,
                        'vendor2_tgl': v2_tgl,
                        'vendor2_resi': v2_resi,
                        'vendor2_biaya': v2_biaya,
                        'keterangan': keterangan,
                        'status': status,
                        'tgl_bayar': tgl_bayar,
                        'nama_bayar': nama_bayar,
                        'profit': profit,
                    }
                )
                count += 1
                self.stdout.write(f"  OK: {no_resi} - {pengirim} - Rp{total:,}")
                
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"  Error: {e}"))
        
        wb.close()
        self.stdout.write(self.style.SUCCESS(f"  ✅ Import {count} Outbound"))

    # ==========================
    # IMPORT PENERIMAAN dari CSV
    # ==========================
    def import_penerimaan_csv(self, filepath):
        import csv
        import re
        
        self.stdout.write(f"\n💰 Import PENERIMAAN dari CSV: {filepath}")
        
        akun_pendapatan, _ = Akun.objects.get_or_create(
            kode='403', 
            defaults={'nama': 'Pendapatan Lain-lain', 'kategori': 'REVENUE'}
        )
        akun_kas, _ = Akun.objects.get_or_create(
            kode='101',
            defaults={'nama': 'Kas', 'kategori': 'ASSET'}
        )
        
        with open(filepath, mode='r', encoding='utf-8-sig') as csv_file:
            reader = csv.reader(csv_file)
            for _ in range(5):
                next(reader, None)
            
            count = 0
            jurnal_count = 0
            
            for row in reader:
                if len(row) < 4:
                    continue
                
                keterangan = row[2].strip() if len(row) > 2 else ''
                if not keterangan:
                    continue

                try:
                    tgl = self.parse_date_str(row[1])
                    nilai = self.parse_rupiah(row[3]) if len(row) > 3 else 0
                    saldo = self.parse_rupiah(row[4]) if len(row) > 4 else 0
                    catatan = row[5].strip() if len(row) > 5 else ''
                    
                    if not tgl or nilai == 0:
                        continue
                    
                    obj, created = Penerimaan.objects.update_or_create(
                        tanggal=tgl,
                        keterangan=keterangan,
                        defaults={
                            'nilai': nilai,
                            'saldo': saldo,
                            'catatan': catatan,
                        }
                    )
                    count += 1
                    
                    uraian = f"Penerimaan: {keterangan}"
                    if not Jurnal.objects.filter(tanggal=tgl, uraian=uraian[:255], nominal=nilai).exists():
                        Jurnal.objects.create(
                            tanggal=tgl,
                            uraian=uraian[:255],
                            akun_debit=akun_kas,
                            akun_kredit=akun_pendapatan,
                            nominal=nilai
                        )
                        jurnal_count += 1
                    
                except Exception as e:
                    self.stdout.write(self.style.ERROR(f"  Error: {e}"))

        self.stdout.write(self.style.SUCCESS(f"  ✅ Import {count} Penerimaan + {jurnal_count} Jurnal"))

    # ==========================
    # IMPORT MANIFEST dari CSV Files (6 Kategori)
    # ==========================
    def import_manifest_csvs(self, base_dir):
        import csv
        
        # Mapping nama file ke kategori dan config kolom
        files_map = [
            {'file': 'Manifest_Bmm_Agustus_2022.csv', 'kategori': 'HULU', 'layout': 'standard'},
            {'file': 'Pantura.csv', 'kategori': 'PANTURA', 'layout': 'standard'},
            {'file': 'PUTUSSIBAU.csv', 'kategori': 'PUTUSSIBAU', 'layout': 'standard'},
            {'file': 'TRUK.csv', 'kategori': 'TRUK', 'layout': 'standard'},
            {'file': 'KALTENG.csv', 'kategori': 'KALTENG', 'layout': 'standard'},
            {'file': 'Ketapang.csv', 'kategori': 'KETAPANG', 'layout': 'shifted'},
        ]

        total_count = 0
        self.stdout.write("\n📦 Import MANIFEST dari 6 File CSV:")

        for item in files_map:
            filepath = os.path.join(base_dir, item['file'])
            if not os.path.exists(filepath):
                self.stdout.write(self.style.WARNING(f"  ❌ File not found: {item['file']}"))
                continue
            
            self.stdout.write(f"  📂 Processing {item['file']} ({item['kategori']})...")
            count = 0
            
            with open(filepath, mode='r', encoding='utf-8-sig', errors='ignore') as csv_file:
                reader = csv.reader(csv_file)
                # Skip header lines (biasanya 5-7 baris)
                # Kita cari baris yang dimulai dengan angka (No urut) di kolom 0
                for row in reader:
                    # Validasi baris data: Kolom 0 angka, minimal 10 kolom
                    if not row or len(row) < 10:
                        continue
                    
                    try:
                        # Cek kolom 0 harus angka integer (No)
                        int(row[0])
                    except:
                        continue # Skip header/garbage row

                    try:
                        layout = item['layout']
                        
                        if layout == 'standard':
                            # Format Standar
                            no_resi = row[2].strip()
                            if not no_resi: continue

                            tgl_kirim = self.parse_date_str(row[1])
                            pengirim = row[3].strip()
                            tujuan = row[4].strip()
                            koli = self.parse_int(row[5])
                            kg = self.parse_float(row[6])
                            penerima = row[7].strip()
                            tgl_terima = self.parse_date_str(row[8]) if len(row) > 8 else None
                            
                            # Logika Baru: Ambil semua angka rupiah setelah kolom Penerima
                            # Cari mulai index 8 sampai akhir
                            tarif, total = self.extract_tarif_total(row, start_idx=8)

                        elif layout == 'shifted':
                            # Format Ketapang
                            no_resi = row[3].strip()
                            if not no_resi: continue

                            tgl_kirim = self.parse_date_str(row[1])
                            tujuan = row[4].strip()
                            koli = self.parse_int(row[5])
                            kg = self.parse_float(row[6])
                            penerima = row[7].strip()
                            pengirim = row[8].strip() if len(row) > 8 else ''
                            tgl_terima = self.parse_date_str(row[9]) if len(row) > 9 else None
                            
                            # Cari mulai index 9 (setelah Tgl Terima/Pengirim)
                            tarif, total = self.extract_tarif_total(row, start_idx=9)

                        # HAPUS LOGIKA AUTO HITUNG (Berbahaya jika salah parse)
                        # if total == 0 and kg > 0 and tarif > 0:
                        #    total = int(kg * tarif)

                        obj, created = Manifest.objects.update_or_create(
                            no_resi=no_resi,
                            kategori=item['kategori'],
                            defaults={
                                'tanggal_kirim': tgl_kirim,
                                'pengirim': pengirim,
                                'tujuan': tujuan,
                                'koli': koli,
                                'kg': kg,
                                'penerima': penerima,
                                'tanggal_terima': tgl_terima,
                                'tarif': tarif,
                                'total': total,
                            }
                        )
                        count += 1

                    except Exception as e:
                        pass
            
            self.stdout.write(f"    ✓ {count} data imported.")
            total_count += count
            
        self.stdout.write(self.style.SUCCESS(f"  ✅ Total Import {total_count} Manifest records from CSVs"))

    def extract_tarif_total(self, row, start_idx):
        """
        Smart extraction untuk Tarif dan Total.
        Mengumpulkan semua nilai rupiah dari start_idx ke kanan.
        - Jika ketemu 2 angka: Angka 1 = Tarif, Angka 2 = Total
        - Jika ketemu 1 angka: Tarif = 0, Angka 1 = Total
        """
        rupiah_values = []
        # Scan sampai max index row
        for i in range(start_idx, len(row)):
            val = self.parse_rupiah(row[i])
            # Hanya ambil jika > 100 perak (asumsi valid price) atau jika formatnya jelas string duit
            if val > 0:
                rupiah_values.append(val)
        
        # Logika Assignment
        if len(rupiah_values) >= 2:
            return rupiah_values[0], rupiah_values[1] # Tarif, Total
        elif len(rupiah_values) == 1:
            return 0, rupiah_values[0] # Tarif Kosong, Total Ada
        else:
            return 0, 0

    # ============ UTILITIES ============
    def parse_excel_date(self, val):
        """Parse date dari Excel (bisa datetime object atau string)"""
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.date()
        return self.parse_date_str(str(val))
    
    def parse_date_str(self, date_str):
        if not date_str or date_str.strip() == '':
            return None
        date_str = date_str.strip()
        
        date_formats = [
            '%m/%d/%Y', '%d/%m/%Y', '%m/%d/%y', '%d/%m/%y',
            '%Y-%m-%d', '%d-%m-%Y'
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except:
                continue
        return None

    def parse_int(self, val):
        if val is None: 
            return 0
        try:
            return int(float(val))
        except:
            return 0

    def parse_float(self, val):
        if val is None: 
            return 0.0
        val_str = str(val).strip()
        if '/' in val_str:
            return 0.0
        try:
            return float(val_str.replace(',', '').replace('"', ''))
        except:
            return 0.0

    def parse_rupiah(self, val):
        """Parse format Rupiah dari CSV/Excel"""
        import re
        if val is None:
            return 0
        if isinstance(val, (int, float)):
            return int(val)
            
        val_str = str(val).strip()
        # Hapus karakter mata uang dan quote. JANGAN hapus koma/titik dulu.
        val_str = re.sub(r'[Rp\s"]', '', val_str)
        
        if not val_str or val_str == '-': 
            return 0
            
        # Deteksi format
        if ',' in val_str and '.' in val_str:
            last_comma = val_str.rfind(',')
            last_dot = val_str.rfind('.')
            
            if last_comma < last_dot: 
                # Format US: 1,500,000.00 -> 1500000.00
                val_str = val_str.replace(',', '')
            else: 
                # Format Indo: 1.500.000,00 -> 1500000.00
                val_str = val_str.replace('.', '').replace(',', '.')
        
        elif ',' in val_str:
            # Kemungkinan besar ribuan (1,500) di format US
            val_str = val_str.replace(',', '')
            
        elif val_str.count('.') > 1:
            # Format 1.500.000 (Indo tanpa desimal)
            val_str = val_str.replace('.', '')
            
        try:
            return int(float(val_str))
        except:
            return 0
